"""
build_aq_reference_v2.py — AQ scoring reference (XLSX + PDF)

v2.0 — complete rewrite. The v1 script carried its own hardcoded copy of the
hole data and collapsed each `Xl`/`Xr` sector pair into a single named row.
That data model can no longer represent the master: 61 of 64 l/r pairs now
carry different score vectors, so collapsing them is lossy. v2 holds NO score
data of its own.

Reads:
  all_holes_data_v3_64.py   — pins, sector_order, scores   (the master)
  build_aqmod_data.py       — DESCRIPTIONS dict            (sector prose)
Writes:
  AQ_Reference_v2.2.xlsx    — summary sheet + 18 per-hole sheets
  AQ_Reference_v2.2.pdf     — 2 holes per page

One row per coded sector, in sector_order (priority order — earlier wins).
"""

import ast
import os
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                Paragraph, Spacer, PageBreak)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# ── CONFIG ──
HERE = Path(__file__).parent
MASTER_PY = HERE / "all_holes_data_v3_64.py"
GENERATOR_PY = HERE / "build_aqmod_data.py"     # source of DESCRIPTIONS
OUT_DIR = Path("/mnt/user-data/outputs")
VERSION = "v2.2"

PAR = {1: 4, 2: 4, 3: 4, 4: 4, 5: 3, 6: 4, 7: 4, 8: 3, 9: 5,
       10: 4, 11: 4, 12: 4, 13: 3, 14: 5, 15: 5, 16: 3, 17: 5, 18: 4}

# Locked per-hole slope reference (from Joel). "R→L" = right is the HIGH side.
SLOPE = {
    1:  "very slight R→L",
    2:  "R→L back, steep R→L front",
    3:  "flat",
    4:  "slight R→L",
    5:  "flat",
    6:  "slight R→L",
    7:  "R→L back, flat front",
    8:  "flat",
    9:  "very slight L→R",
    10: "flat (back 2/3 has a B→F + left component)",
    11: "flat",
    12: "flat",
    13: "slight R→L front, L→R back (opposite directions)",
    14: "flat front, L→R back",
    15: "severe R→L front, slight R→L back",
    16: "slight R→L on top quarter only (S6); S4/S5 steep B→F, no L/R",
    17: "slight L→R (entire green)",
    18: "flat",
}


# ── LOAD LIVE DATA ──

def load_master(path):
    ns = {}
    exec(open(path).read(), ns)
    return ns["HOLES"]


def load_descriptions(path):
    """Parse the DESCRIPTIONS literal without executing the generator
    (running it would rewrite holes_data.js as a side effect)."""
    tree = ast.parse(open(path).read())
    for node in tree.body:
        if isinstance(node, ast.Assign) and getattr(node.targets[0], "id", "") == "DESCRIPTIONS":
            return ast.literal_eval(node.value)
    raise RuntimeError(f"DESCRIPTIONS not found in {path}")


HOLES = load_master(MASTER_PY)
DESCRIPTIONS = load_descriptions(GENERATOR_PY)

# Fail loudly rather than emit a reference with blank rows
for h in range(1, 19):
    missing = set(HOLES[h]["sector_order"]) - set(DESCRIPTIONS.get(h, {}))
    if missing:
        raise RuntimeError(f"H{h}: no description for sector(s) {sorted(missing)}")


def rows_for(h):
    """[(row_label, [5 scores]), ...] in sector_order."""
    d = HOLES[h]
    return [(f"S{sk} — {DESCRIPTIONS[h][sk]}", list(d["scores"][sk]))
            for sk in d["sector_order"]]


def pins_for(h):
    return [HOLES[h]["pin_labels"][i] for i in range(1, 6)]


def header_for(h):
    return f"Hole {h}  (par {PAR[h]}, {len(HOLES[h]['sector_order'])} sectors)"


# ── SHARED COLOR SCALE ──

BANDS = [(4.5, '1B7A2B'), (4.0, '4CAF50'), (3.5, '8BC34A'), (3.0, 'CDDC39'),
         (2.5, 'FFC107'), (2.0, 'FF9800'), (1.5, 'FF5722')]
WORST = 'B71C1C'


def score_hex(v):
    for thresh, hexc in BANDS:
        if v >= thresh:
            return hexc
    return WORST


def light_text(v):
    return v >= 4.5 or v <= 1.5


# ── XLSX ──

header_fill = PatternFill('solid', fgColor='2F5233')
header_font_white = Font(bold=True, size=11, name='Arial', color='FFFFFF')
sector_font = Font(size=10, name='Arial')
hole_title_font = Font(bold=True, size=14, name='Arial')
slope_font = Font(italic=True, size=10, name='Arial', color='555555')
thin = Side(style='thin')
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)


def score_cell(c, v):
    c.fill = PatternFill('solid', fgColor=score_hex(v))
    c.font = Font(size=11, name='Arial', bold=True,
                  color='FFFFFF' if light_text(v) else '000000')
    c.border = thin_border
    c.alignment = Alignment(horizontal='center')
    c.number_format = '0.0'


def write_block(ws, h, row):
    ws.cell(row=row, column=1, value=header_for(h)).font = hole_title_font
    row += 1
    ws.cell(row=row, column=1, value=f'Slope: {SLOPE[h]}').font = slope_font
    row += 1

    c = ws.cell(row=row, column=1, value='Sector')
    c.font, c.fill, c.border = header_font_white, header_fill, thin_border
    for j, pin in enumerate(pins_for(h)):
        c = ws.cell(row=row, column=j + 2, value=f'P{j+1} ({pin})')
        c.font, c.fill, c.border = header_font_white, header_fill, thin_border
        c.alignment = Alignment(horizontal='center')
    row += 1

    for label, scores in rows_for(h):
        c = ws.cell(row=row, column=1, value=label)
        c.font, c.border = sector_font, thin_border
        for j, v in enumerate(scores):
            score_cell(ws.cell(row=row, column=j + 2, value=v), v)
        row += 1
    return row


def widths(ws):
    ws.column_dimensions['A'].width = 40
    for col in range(2, 7):
        ws.column_dimensions[get_column_letter(col)].width = 12


wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'All Holes'
r = 1
for h in range(1, 19):
    r = write_block(ws, h, r) + 2
widths(ws)

for h in range(1, 19):
    ws2 = wb.create_sheet(title=f'Hole {h}')
    write_block(ws2, h, 1)
    widths(ws2)

OUT_DIR.mkdir(parents=True, exist_ok=True)
xlsx_path = OUT_DIR / f'AQ_Reference_{VERSION}.xlsx'
wb.save(xlsx_path)
print(f'XLSX: {xlsx_path}')


# ── PDF ──

styles = getSampleStyleSheet()
title_style = ParagraphStyle('HoleTitle', parent=styles['Heading2'],
                             fontSize=13, spaceAfter=2, spaceBefore=0)
slope_style = ParagraphStyle('Slope', parent=styles['Normal'], fontSize=9,
                             textColor=colors.grey, spaceAfter=4)

doc = SimpleDocTemplate(str(OUT_DIR / f'AQ_Reference_{VERSION}.pdf'),
                        pagesize=letter, topMargin=0.5 * inch,
                        bottomMargin=0.5 * inch, leftMargin=0.5 * inch,
                        rightMargin=0.5 * inch)

story = [
    Paragraph('Oberlin Golf Club', styles['Title']),
    Paragraph(f'AQ Scoring Reference — {VERSION}', styles['Heading2']),
    Spacer(1, 12),
    Paragraph(f'Generated from <b>{MASTER_PY.name}</b>. One row per coded sector, '
              'listed in <i>sector_order</i> (priority order — earlier wins on overlap).',
              styles['Normal']),
    Spacer(1, 8),
    Paragraph('<b>Scale:</b> 5 = best realistic outcome, 1 = disaster (half-points used)', styles['Normal']),
    Paragraph('<b>Approach window:</b> 40–180 yards; non-qualifying holes are NA', styles['Normal']),
    Paragraph('<b>Round AQ:</b> average of qualifying approach scores', styles['Normal']),
    Paragraph('<b>Pin positions:</b> 1–5 per pin sheet, categorical (not a ranking)', styles['Normal']),
    Paragraph('<b>Slope grammar:</b> "R→L" means right is the HIGH side; the ball runs right to left.',
              styles['Normal']),
    PageBreak(),
]

for idx, h in enumerate(range(1, 19)):
    story.append(Paragraph(header_for(h), title_style))
    story.append(Paragraph(f'Slope: {SLOPE[h]}', slope_style))

    table_data = [['Sector'] + [f'P{j+1} ({p})' for j, p in enumerate(pins_for(h))]]
    for label, scores in rows_for(h):
        table_data.append([label] + [f'{v:.1f}' for v in scores])

    t = Table(table_data, colWidths=[3.0 * inch] + [0.8 * inch] * 5, repeatRows=1)
    cmds = [
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 7.5),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2F5233')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, -1), 1.5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1.5),
    ]
    for i, (_, scores) in enumerate(rows_for(h)):
        for j, v in enumerate(scores):
            cmds += [
                ('BACKGROUND', (j + 1, i + 1), (j + 1, i + 1), colors.HexColor('#' + score_hex(v))),
                ('TEXTCOLOR', (j + 1, i + 1), (j + 1, i + 1),
                 colors.white if light_text(v) else colors.black),
                ('FONTNAME', (j + 1, i + 1), (j + 1, i + 1), 'Helvetica-Bold'),
            ]
    t.setStyle(TableStyle(cmds))
    story.append(t)
    story.append(Spacer(1, 14))
    if idx % 2 == 1 and h != 18:
        story.append(PageBreak())

doc.build(story)
print(f'PDF:  {OUT_DIR / f"AQ_Reference_{VERSION}.pdf"}')
print(f'Holes: {list(range(1, 19))}  |  rows: {sum(len(HOLES[h]["sector_order"]) for h in range(1, 19))}')
